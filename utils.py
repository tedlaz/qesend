from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from model_data import ModelData


def data2csv(data: ModelData, filename: str, separator: str, encoding: str):
    txt = separator.join(data.headers) + "\n"
    for line in data.rows:
        txt += separator.join([str(i) for i in line]) + "\n"
    with open(filename, "w", encoding=encoding) as fil:
        fil.write(txt)


def grnum(num):
    if num == 0:
        return ""
    return f"{num:,.2f}".replace(",", "|").replace(".", ",").replace("|", ".")


def data2xlsx(data: ModelData, filename: str):
    wb1 = Workbook()
    fbold = Font(bold=True)
    acenter = Alignment(horizontal="center", wrapText=True)
    # wb1.iso_dates = True
    ws1 = wb1.active
    ws1.title = "data"
    for i, title in enumerate(data.headers):
        ws1.cell(row=1, column=i + 1, value=title)
        ws1.cell(row=1, column=i + 1).font = fbold
        ws1.cell(row=1, column=i + 1).alignment = acenter

    for i, line in enumerate(data.rows):
        crow = i + 2
        for j, val in enumerate(line):
            if data.typos[j] == 4:
                ws1.cell(row=crow, column=j + 1, value=date.fromisoformat(val))
                ws1.cell(crow, j + 1).number_format = "DD/MM/YYYY"
                continue
            if data.typos[j] == 3:
                ws1.cell(row=crow, column=j + 1, value=val)
                ws1.cell(crow, j + 1).number_format = "#,##0.00"
                continue
            ws1.cell(row=crow, column=j + 1, value=val)

    wb1.save(filename)
